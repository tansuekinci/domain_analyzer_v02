import socket
import requests
import whois
import sys
import re
import time
import os
import pandas as pd
import dns.resolver
import concurrent.futures
import random
from pathlib import Path 
from tabulate import tabulate
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Border, Side, PatternFill
import subprocess 

# --- GLOBAL CONFIGURATION ---
PSI_API_KEY = "Google_API_KEY" 
# AbuseIPDB API Key'inizi buraya girin. Boş bırakılırsa simülasyon çalışır.
ABUSEIPDB_API_KEY = "ABUSEIPDB_API_KEY" 

# --- CONSOLE STYLING ---
GREEN = "\033[92m"
RED = "\033[91m"
ENDC = "\033[0m"
CHECK = f"{GREEN}✅{ENDC}"
CROSS = f"{RED}❌{ENDC}"
RBL_LIST = [
    'zen.spamhaus.org', 
    'bl.spamcop.net',
    'b.barracudacentral.org',
]

ROTATING_USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2 Safari/605.1.15',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:109.0) Gecko/20100101 Firefox/121.0'
]

COMMON_PORTS = {
    21: 'FTP', 22: 'SSH', 23: 'Telnet', 25: 'SMTP', 53: 'DNS',
    80: 'HTTP', 110: 'POP3', 143: 'IMAP', 443: 'HTTPS', 465: 'SMTPS',
    587: 'SMTP (Sub)', 993: 'IMAPS', 1433: 'MSSQL', 3306: 'MySQL',
    3389: 'RDP', 5432: 'PostgreSQL', 8080: 'HTTP-Proxy', 8443: 'HTTPS-Alt'
}

class SuppressStderr:
    def __enter__(self):
        self._original_stderr = sys.stderr
        self._null = open(os.devnull, 'w')
        sys.stderr = self._null
    def __exit__(self, exc_type, exc_val, exc_tb):
        sys.stderr = self._original_stderr
        self._null.close()

def is_valid_ip(text):
    pattern = r"^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
    return re.match(pattern, text) is not None

def clean_val(v):
    if v is None: return "-"
    if isinstance(v, list): return ", ".join(str(x) for x in v if x is not None)
    return str(v)

def get_ip_details(ip):
    try:
        if not is_valid_ip(ip): return None
        time.sleep(0.3)
        current_headers = {'User-Agent': random.choice(ROTATING_USER_AGENTS)}
        response = requests.get(f"http://ip-api.com/json/{ip}?fields=status,country,city,isp,org,as", headers=current_headers, timeout=3)
        return response.json()
    except: return None

def scan_single_port(ip, port, service_name):
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(0.8)
        result = s.connect_ex((ip, port))
        if result == 0:
            banner = "Unknown"
            try:
                if port in [80, 8080, 443]: s.send(b"HEAD / HTTP/1.1\r\n\r\n")
                banner = s.recv(1024).decode('utf-8', errors='ignore').strip()[:50].replace('\r', '').replace('\n', ' ')
            except: pass
            s.close()
            return {'ip': ip, 'port': port, 'service': service_name, 'status': 'OPEN', 'banner': banner}
        s.close()
        return None
    except: return None

def scan_target_services(ip_list):
    scan_results = []
    ips_to_scan = list(ip_list)[:10] 
    if not ips_to_scan: return []
        
    print(f"   > Scanning {len(ips_to_scan)} unique IPs (First 10)...")

    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
        futures = []
        for ip in ips_to_scan:
            for port, name in COMMON_PORTS.items():
                futures.append(executor.submit(scan_single_port, ip, port, name))
        for future in concurrent.futures.as_completed(futures):
            res = future.result()
            if res: scan_results.append(res)
    return scan_results

def analyze_pagespeed(target_url):
    api_url = f"https://www.googleapis.com/pagespeedonline/v5/runPagespeed?url={target_url}&strategy=desktop&category=performance&category=seo&category=accessibility"
    if PSI_API_KEY: api_url += f"&key={PSI_API_KEY}"

    try:
        response = requests.get(api_url, timeout=60)
        data = response.json()
        if 'error' in data: return None, [f"API Error: {data['error']['message']}"]

        lighthouse = data.get('lighthouseResult', {})
        categories = lighthouse.get('categories', {})
        audits = lighthouse.get('audits', {})

        scores = {
            'Performance': int(categories.get('performance', {}).get('score', 0) * 100),
            'Accessibility': int(categories.get('accessibility', {}).get('score', 0) * 100),
            'SEO': int(categories.get('seo', {}).get('score', 0) * 100)
        }
        metrics = {
            'FCP': audits.get('first-contentful-paint', {}).get('displayValue', '-'),
            'LCP': audits.get('largest-contentful-paint', {}).get('displayValue', '-'),
            'Speed Idx': audits.get('speed-index', {}).get('displayValue', '-')
        }
        recommendations = []
        for k, v in audits.items():
            if v.get('details', {}).get('type') == 'opportunity' and v.get('score', 1) < 0.9:
                recommendations.append([v.get('title'), v.get('displayValue', '')])
        
        return {
            'url': target_url,
            'scores': scores,
            'metrics': metrics,
            'recommendations': recommendations
        }, None

    except Exception as e: return None, [f"Analysis failed for {target_url}: {str(e)}"]

def get_dns_records_robust(domain):
    dns_results = []
    record_types = ['A', 'MX', 'NS', 'TXT', 'SOA', 'CNAME']
    # 1. Google DoH
    try:
        for r_type in record_types:
            resp = requests.get("https://dns.google/resolve", params={'name': domain, 'type': r_type}, timeout=5)
            if resp.status_code == 200 and 'Answer' in resp.json():
                for ans in resp.json()['Answer']:
                    dns_results.append({'type': r_type, 'value': ans['data'].replace('"', '')})
    except: pass
    
    # 2. Cloudflare Fallback
    if not dns_results: 
        try:
            cf_headers = {'Accept': 'application/dns-json'}
            for r_type in record_types:
                resp = requests.get(f"https://cloudflare-dns.com/dns-query?name={domain}&type={r_type}", headers=cf_headers, timeout=5)
                if resp.status_code == 200 and 'Answer' in resp.json():
                    for ans in resp.json()['Answer']:
                        dns_results.append({'type': r_type, 'value': ans['data'].replace('"', '')})
        except: pass
    
    return dns_results

def discover_subdomains_sublist3r(domain):
    base_dir = Path(__file__).resolve().parent 
    sublist3r_path = base_dir / "Sublist3r" / "sublist3r.py"
    output_file_path = base_dir / "Sublist3r" / f"{domain}.txt"
    
    subdomain_results = []
    MAX_ATTEMPTS = 8 
    WAIT_SECONDS = 60 

    if not sublist3r_path.exists():
        print(f"   [!] ERROR: Sublist3r script not found: {sublist3r_path}")
        return []

    print(f"   > Running Sublist3r (up to {MAX_ATTEMPTS} attempts). Output path: {output_file_path.name}")
    command = [
        "python3", str(sublist3r_path), "-d", domain, "-e", "ssl,passivedns", "-o", str(output_file_path),
    ]
    
    success = False
    
    for attempt in range(1, MAX_ATTEMPTS + 1):
        
        print(f"   > Attempt #{attempt}/{MAX_ATTEMPTS}...")
        
        if output_file_path.exists():
            try: os.remove(output_file_path)
            except: pass
            
        try:
            with open(os.devnull, 'w') as DEVNULL:
                process = subprocess.run(
                    command, text=True, timeout=180,
                    cwd=str(sublist3r_path.parent),
                    stdout=DEVNULL, stderr=subprocess.PIPE 
                ) 
            
            if process.returncode != 0:
                print(f"   [!] Sublist3r failed (Attempt {attempt}). Error Code: {process.returncode}")
            
            if output_file_path.exists() and output_file_path.stat().st_size > 0:
                print(f"   [SUCCESS] Sublist3r output file created and is not empty on attempt #{attempt}.")
                success = True
                break 
            
            elif output_file_path.exists() and output_file_path.stat().st_size == 0:
                 print(f"   [!] Output file created but is EMPTY on attempt #{attempt}. Retrying...")
            
            else: 
                 print(f"   [!] Output file NOT created on attempt #{attempt}. Retrying...")

        except subprocess.TimeoutExpired:
            print(f"   [!] ERROR: Sublist3r timed out (Attempt {attempt}). Retrying...")
        except FileNotFoundError:
            print("   [!] ERROR: python3 command not found. Check your PATH settings. Aborting retries.")
            return [] 
        except Exception as e:
            print(f"   [!] Unexpected error during execution (Attempt {attempt}): {e}. Retrying...")

        if not success and attempt < MAX_ATTEMPTS:
            print(f"   > Waiting {WAIT_SECONDS} seconds before next attempt (User requested).")
            time.sleep(WAIT_SECONDS)

    if not success:
        print(f"   [FAILED] Sublist3r failed to generate valid output after {MAX_ATTEMPTS} attempts. Aborting subdomain resolution.")
        
    if success:
        try:
            with open(output_file_path, 'r') as f:
                subdomains = [line.strip() for line in f if line.strip()]
            
            print(f"   > Resolving {len(subdomains)} subdomains to IP addresses...")
            
            for sub in subdomains:
                record_type = "N/A"; ip = "Failed to Resolve"
                
                try:
                    ip = socket.gethostbyname(sub); record_type = 'A'
                except socket.gaierror:
                    try:
                        resp = requests.get("https://dns.google/resolve", params={'name': sub, 'type': 'CNAME'}, timeout=3)
                        if resp.status_code == 200 and 'Answer' in resp.json():
                            ip = resp.json()['Answer'][0]['data'].replace('"', ''); record_type = 'CNAME'
                    except: pass
                
                subdomain_results.append({
                    'Subdomain': sub, 'Type': record_type, 'Value/IP': ip, 'Last_Seen': 'Sublist3r'
                })
        
        except Exception as e:
            print(f"   [!] ERROR: Failed to read or resolve output file: {e}")
            subdomain_results = []
        finally:
            if output_file_path.exists(): os.remove(output_file_path)
            
    return subdomain_results

def get_rapiddns_history(domain):
    history = []
    current_headers = {'User-Agent': random.choice(ROTATING_USER_AGENTS)}
    try:
        r = requests.get(f"https://rapiddns.io/s/{domain}", headers=current_headers, timeout=10)
        if r.status_code!=200: return []
        soup = BeautifulSoup(r.text, 'html.parser')
        table = soup.find('table', {'class':'table'})
        if not table: return []
        for row in table.find_all('tr')[1:]:
            cols = row.find_all('td')
            f_ip, f_date, f_type = None, "-", "Unknown"
            for c in cols:
                txt = c.text.strip()
                if is_valid_ip(txt): f_ip=txt
                elif re.match(r"\d{4}-\d{2}-\d{2}", txt): f_date=txt
                elif txt in ['A','MX','CNAME']: f_type=txt
            if f_type=='A' and f_ip: history.append({'ip':f_ip, 'date':f_date})
        return history
    except: return []

def write_domain_info_to_files(domain, root_ip, root_dns_records, subdomain_data, scan_data, write_dns=True, write_webservers=True):
    base_dir = Path(__file__).resolve().parent 
    
    webservers_file_path = base_dir / f"{domain}_webservers.txt"
    
    if write_dns:
        # --- 1. DNS File (domainadi_dns.txt) ---
        dns_file_path = base_dir / f"{domain}_dns.txt"
        dns_lines = set()

        dns_lines.add(f"Root Domain: {domain} | IP: {root_ip}")
        for rec in root_dns_records: dns_lines.add(f"{rec['VALUE']} | {rec['TYPE']}") 
        for sub in subdomain_data: dns_lines.add(f"{sub['Subdomain']} | {sub['Type']} | {sub['Value/IP']}")

        try:
            with open(dns_file_path, 'w', encoding='utf-8') as f:
                f.write(f"# {domain.upper()} - All DNS Records (Root and Subdomains)\n")
                for line in sorted(list(dns_lines)): f.write(line + "\n")
            print(f"   [+] DNS records saved to file: {dns_file_path.name}")
        except Exception as e: print(f"   [!] ERROR: Failed to write DNS records: {e}")
        
    if write_webservers:
        # --- 2. Webservers File (domainadi_webservers.txt) ---
        webserver_lines = set()
        
        ip_to_domain_map = {}
        if is_valid_ip(root_ip): ip_to_domain_map[root_ip] = {domain}
        
        for sub in subdomain_data:
            ip = sub['Value/IP']
            if is_valid_ip(ip):
                if ip not in ip_to_domain_map: ip_to_domain_map[ip] = set()
                ip_to_domain_map[ip].add(sub['Subdomain'])

        open_web_ips = set()
        for scan in scan_data:
            if isinstance(scan, dict) and 'port' in scan:
                port = scan['port']
                if port == 80 or port == 443: open_web_ips.add(scan['ip'])
                
        for ip in open_web_ips:
            domains = sorted(list(ip_to_domain_map.get(ip, {f"Unknown domain resolving to IP ({domain})" if ip == root_ip else ip})))
            for d in domains: webserver_lines.add(f"{d} | {ip}")
        
        try:
            with open(webservers_file_path, 'w', encoding='utf-8') as f:
                f.write(f"# {domain.upper()} - Domain/Subdomain and IP Matches with Open 80/443 Ports\n")
                for line in sorted(list(webserver_lines)): f.write(line + "\n")
            print(f"   [+] Webserver records saved to file: {webservers_file_path.name}")
            return webservers_file_path
        except Exception as e:
            print(f"   [!] ERROR: Failed to write Webserver records: {e}")
            return None
    
    return webservers_file_path

def read_dns_data_from_file(domain, root_ip):
    base_dir = Path(__file__).resolve().parent
    dns_file_path = base_dir / f"{domain}_dns.txt"
    
    subdomain_data_excel = []
    root_dns_records_for_file = []
    dns_display_list = []
    scan_ips = set()
    
    try:
        with open(dns_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'): continue
                
                parts = line.split('|')
                
                if line.startswith('Root Domain:'):
                    ip_part = parts[1].split(':')[-1].strip()
                    if is_valid_ip(ip_part): scan_ips.add(ip_part)
                elif len(parts) == 2:
                    val = parts[0].strip()
                    type_rec = parts[1].strip()
                    root_dns_records_for_file.append({'TYPE': type_rec, 'VALUE': val})
                    dns_display_list.append([type_rec, val])
                    if type_rec == 'A' and is_valid_ip(val): scan_ips.add(val)
                elif len(parts) == 3:
                    subdomain = parts[0].strip()
                    type_rec = parts[1].strip()
                    val = parts[2].strip()
                    subdomain_data_excel.append({'Subdomain': subdomain, 'Type': type_rec, 'Value/IP': val, 'Last_Seen': 'File'})
                    if type_rec == 'A' and is_valid_ip(val): scan_ips.add(val)

        print(f"   [+] Existing DNS records successfully loaded from '{dns_file_path.name}'.")
        
        if is_valid_ip(root_ip): scan_ips.add(root_ip)
        
        # Console output data
        subdomain_data_console = [[s['Subdomain'], s['Type'], s['Value/IP'], s['Last_Seen']] for s in subdomain_data_excel]
        
        return subdomain_data_excel, root_dns_records_for_file, dns_display_list, scan_ips, subdomain_data_console 
        
    except Exception as e:
        print(f"   [!] ERROR: An error occurred while reading the existing DNS file: {e}")
        return [], [], [], set(), []


def perform_discovery_tasks(domain, c_ip):
    """Performs DNS and Subdomain discovery, collects IPs, and returns console/file lists."""
    scan_ips = set() 
    if is_valid_ip(c_ip): scan_ips.add(c_ip) 
    
    # 2. DNS RECORDS (ROOT DOMAIN)
    print("⏳ Querying DNS Records...")
    recs = get_dns_records_robust(domain)
    root_dns_records_for_file = []
    dns_display_list = []
    if recs:
        unique = []; seen = set()
        for r in recs:
            if (r['type'], r['value']) not in seen:
                seen.add((r['type'], r['value'])); unique.append(r)
                if r['type']=='A' and is_valid_ip(r['value']): scan_ips.add(r['value']) 
            root_dns_records_for_file.append({'TYPE': r['type'], 'VALUE': r['value']})
        for r in unique:
            val = r['value'];
            if len(val) > 80: val = val[:77]+"..."
            dns_display_list.append([r['type'], val])
    else: dns_display_list.append(["DNS", "No records found"])

    # 3. SUBDOMAIN ENUMERATION 
    print("⏳ Running Subdomain Enumeration (via Sublist3r)...")
    subdomains_found = discover_subdomains_sublist3r(domain)
    subdomain_data_console = []
    subdomain_data_excel = []
    
    if subdomains_found:
        print(f"⚡ Found {len(subdomains_found)} unique subdomains...")
        subdomain_data_excel = subdomains_found 
        subdomain_data_console = [[s['Subdomain'], s['Type'], s['Value/IP'], s['Last_Seen']] for s in subdomains_found]
        
        for item in subdomains_found:
            ip = item.get('Value/IP')
            if is_valid_ip(ip): scan_ips.add(ip)
    else:
        subdomain_data_console.append(["No subdomains found", "-", "-", "Sublist3r Failed or Found None"])
        
    return subdomain_data_excel, root_dns_records_for_file, dns_display_list, scan_ips, subdomain_data_console


def run_portscan_tasks(scan_ips):
    """Performs port scan and returns results."""
    print(f"🚀 Scanning {len(scan_ips)} IP(s) from Root DNS and Subdomains...")
    if scan_ips:
        res = scan_target_services(list(scan_ips))
        if res:
            return res # scan_data (list of dicts)
        else: return ["Result", "No open ports found on scanned IPs", "-"]
    else: return ["Result", "No IP to scan", "-"]
    
# --- MAIL TEST FONKSİYONLARI ---

def get_mx_targets(root_dns_records_for_file):
    """Extracts unique MX records from DNS records."""
    mx_records = []
    for rec in root_dns_records_for_file:
        if rec['TYPE'] == 'MX':
            # MX record VALUE is typically the FQDN ending with a dot
            mx_domain = rec['VALUE'].strip().strip('.')
            mx_records.append(mx_domain)
    # Get unique list of domains/FQDNs to check
    return list(set(mx_records))

def check_mxtoolbox_smtp(target):
    """
    Simulates checking mxtoolbox.com for SMTP health.
    """
    full_url = f"https://mxtoolbox.com/smtp.aspx?action=smtp:{target}"
    time.sleep(1) # Be polite

    # Simulate SMTP Health Check
    results = {
        'Reverse_DNS': 'OK',
        'SMTP_Banner': 'OK',
        'Relay_Check': 'OK',
        'Total_Errors': random.choice([0, 0, 0, 1])
    }
    if results['Total_Errors'] > 0:
        results['Reverse_DNS'] = random.choice(['FAIL', 'OK'])
        
    status = 'PASS' if results['Total_Errors'] == 0 else 'FAIL'
    
    return {
        'Target': target,
        'Test': 'SMTP Health',
        'Status': status,
        'Details': f"RevDNS: {results['Reverse_DNS']}, Banner: {results['SMTP_Banner']}, Errors: {results['Total_Errors']}",
        'Source': "MXToolbox (Simulation)",
        'URL': full_url
    }

def check_dnsbl(target_ip):
    """Checks the given IP against a list of RBLs using dnspython."""
    if not is_valid_ip(target_ip):
        return {'Target': target_ip, 'Test': 'RBL Check', 'Status': 'SKIP', 'Details': 'Invalid IP Address.', 'Source': 'N/A'}

    listed_on = []
    
    # Reverse the IP address: 1.2.3.4 -> 4.3.2.1
    reverse_ip = ".".join(target_ip.split('.')[::-1])
    
    # Check each RBL
    for rbl in RBL_LIST:
        query = f"{reverse_ip}.{rbl}"
        try:
            # Query for an A record. If found, the IP is listed.
            dns.resolver.resolve(query, 'A', lifetime=2.0)
            listed_on.append(rbl)
        except dns.resolver.NXDOMAIN:
            continue
        except dns.resolver.Timeout:
            listed_on.append(f"{rbl} (Timeout/Error)")
        except Exception:
             continue

    if listed_on:
        return {
            'Target': target_ip, 
            'Test': 'RBL Check', 
            'Status': 'LISTED', 
            'Details': f"Listed on: {', '.join(listed_on)}",
            'Source': "Direct DNSBL"
        }
    else:
        return {
            'Target': target_ip, 
            'Test': 'RBL Check', 
            'Status': 'CLEAN', 
            'Details': f"Checked against {len(RBL_LIST)} blacklists. No listing found.",
            'Source': "Direct DNSBL"
        }

def check_abuseipdb(target_ip):
    """Checks the given IP against AbuseIPDB (API or simulation)."""
    if not is_valid_ip(target_ip):
        return {'Target': target_ip, 'Test': 'AbuseIPDB', 'Status': 'SKIP', 'Details': 'Invalid IP Address.', 'Source': 'N/A'}

    if not ABUSEIPDB_API_KEY:
        # Fallback to simulation if no key
        status = random.choice(['CLEAN', 'CLEAN', 'CLEAN', 'LISTED'])
        confidence = random.randint(0, 100) if status == 'LISTED' else 0
        
        return {
            'Target': target_ip, 
            'Test': 'AbuseIPDB', 
            'Status': status, 
            'Details': f"Simulated confidence score: {confidence}% (API Key Missing)",
            'Source': "AbuseIPDB (Simulation)"
        }
        
    # --- Actual API Call (Only if key exists) ---
    api_url = f"https://api.abuseipdb.com/api/v2/check"
    headers = {
        'Accept': 'application/json',
        'Key': ABUSEIPDB_API_KEY 
    }
    params = {
        'ipAddress': target_ip,
        'maxAgeInDays': 90
    }
    
    try:
        response = requests.get(api_url, headers=headers, params=params, timeout=5)
        data = response.json().get('data', {})
        
        if response.status_code == 200 and data:
            score = data.get('abuseConfidenceScore', 0)
            status = 'LISTED' if score > 0 else 'CLEAN'
            return {
                'Target': target_ip, 
                'Test': 'AbuseIPDB', 
                'Status': status, 
                'Details': f"Abuse Confidence Score: {score}% (Total Reports: {data.get('totalReports', 0)})",
                'Source': "AbuseIPDB (API)"
            }
        
        error = response.json().get('errors', [{}])[0].get('detail', 'Unknown API Error')
        return {
            'Target': target_ip, 
            'Test': 'AbuseIPDB', 
            'Status': 'API_ERROR', 
            'Details': f"API failed: {error}",
            'Source': "AbuseIPDB (API)"
        }
        
    except Exception as e:
        return {
            'Target': target_ip, 
            'Test': 'AbuseIPDB', 
            'Status': 'FAIL', 
            'Details': f"Connection Error: {str(e)}",
            'Source': "AbuseIPDB (API)"
        }

def get_dns_txt_record(subdomain):
    """Queries for a specific DNS TXT record using Google DNS."""
    try:
        resp = requests.get("https://dns.google/resolve", params={'name': subdomain, 'type': 'TXT'}, timeout=5)
            
        if resp.status_code == 200 and 'Answer' in resp.json():
            data = " ".join([ans['data'].strip('"') for ans in resp.json()['Answer']])
            return data
    except:
        pass
    return None

def check_email_security_txt(domain):
    """Checks for SPF, DMARC, and common DKIM records."""
    results = []
    
    # --- 1. SPF Check (Root Domain TXT) ---
    spf_record = get_dns_txt_record(domain)
    
    spf_status = 'MISSING'
    spf_details = 'Record not found.'
    
    if spf_record:
        if 'v=spf1' in spf_record.lower():
            spf_status = 'FOUND (Valid)'
            spf_details = spf_record
        else:
            spf_status = 'NOT FOUND (Other TXT)'
            spf_details = f"TXT record found, but v=spf1 tag is missing/invalid: {spf_record[:50]}..."
            
    results.append({'Target': domain, 'Test': 'SPF', 'Status': spf_status, 'Details': spf_details, 'Source': 'Google DNS'})
    
    # --- 2. DMARC Check ---
    dmarc_domain = f'_dmarc.{domain}'
    dmarc_record = get_dns_txt_record(dmarc_domain)
    
    dmarc_status = 'MISSING'
    dmarc_details = 'Record not found.'
    
    if dmarc_record:
        if dmarc_record.startswith('v=DMARC1'):
            dmarc_status = 'FOUND (Valid)'
            dmarc_details = dmarc_record
        else:
            dmarc_status = 'INVALID'
            dmarc_details = f"TXT record found but does not start with v=DMARC1: {dmarc_record[:50]}..."
            
    results.append({'Target': dmarc_domain, 'Test': 'DMARC', 'Status': dmarc_status, 'Details': dmarc_details, 'Source': 'Google DNS'})
    
    # --- 3. DKIM Check ---
    # Attempt to guess common selectors
    common_selectors = ['default', 'selector1', 'google', 'm1']
    dkim_found = False
    
    for selector in common_selectors:
        dkim_domain = f'{selector}._domainkey.{domain}'
        dkim_record = get_dns_txt_record(dkim_domain)
        
        if dkim_record:
            if dkim_record.startswith('v=DKIM1'):
                results.append({'Target': dkim_domain, 'Test': 'DKIM', 'Status': 'FOUND (Valid)', 'Details': dkim_record[:80] + '...', 'Source': 'Google DNS'})
                dkim_found = True
            else:
                results.append({'Target': dkim_domain, 'Test': 'DKIM', 'Status': 'INVALID', 'Details': f"TXT record found but not v=DKIM1: {dkim_record[:50]}...", 'Source': 'Google DNS'})
                dkim_found = True

    if not dkim_found:
        results.append({'Target': f'common selectors (e.g., default.{domain})', 'Test': 'DKIM', 'Status': 'MISSING', 'Details': 'No common DKIM selectors found.', 'Source': 'Google DNS'})
        
    return results

def run_mail_tests(domain, root_dns_records_for_file, mail_choice, root_ip):
    """Main function to run the selected mail tests."""
    mail_test_results = []
    
    run_smtp_health = (mail_choice == '1' or mail_choice == '3')
    run_reputation = (mail_choice == '2' or mail_choice == '3') 
    run_email_auth = (mail_choice == '4' or mail_choice == '3')

    if run_smtp_health:
        # SMTP Health (Requires MX records)
        mx_targets = get_mx_targets(root_dns_records_for_file)
        
        if not mx_targets:
            print(f"\n[!] WARNING: No MX records found for {domain} in the loaded data. SMTP Health test skipped.")
        else:
            print(f"\n[i] Found {len(mx_targets)} MX targets for SMTP testing: {', '.join(mx_targets)}")
            
            for target in mx_targets:
                print(f"   > Running SMTP Health Check for {target}...")
                result = check_mxtoolbox_smtp(target)
                mail_test_results.append(result)
        
    # --- REPUTATION CHECKS (DNSBL / AbuseIPDB) ---
    if run_reputation:
        print(f"\n   > Running Reputation Checks for root domain IP: {root_ip}...")
        
        if is_valid_ip(root_ip):
            
            # 1. DNSBL (Spamhaus, Spamcop, etc.)
            result_dnsbl = check_dnsbl(root_ip) 
            mail_test_results.append(result_dnsbl)
            
            # 2. AbuseIPDB
            result_abuseipdb = check_abuseipdb(root_ip)
            mail_test_results.append(result_abuseipdb)

        else:
            mail_test_results.append({'Target': domain, 'Test': 'Reputation Check', 'Status': 'SKIP', 'Details': 'Root IP not resolved or invalid.', 'Source': 'N/A'})
            
    if run_email_auth:
        # Email Auth Checks (SPF, DMARC, DKIM)
        print(f"\n   > Running Email Authentication Checks (SPF, DMARC, DKIM) for {domain}...")
        email_auth_results = check_email_security_txt(domain)
        mail_test_results.extend(email_auth_results)
        
    return mail_test_results

# --- STYLING FUNCTIONS (No change) ---

def style_subdomain_sheet(ws):
    ws.column_dimensions['A'].width = 35 
    ws.column_dimensions['B'].width = 15 
    ws.column_dimensions['C'].width = 35 
    ws.column_dimensions['D'].width = 15 

def style_history_sheet(ws):
    thick = Side(border_style="thick", color="000000")
    start = 2
    max_r = ws.max_row
    for r in range(2, max_r+2):
        val = ws.cell(r,1).value
        if (val and str(val).startswith("--- RECORD")) or r > max_r:
            if r > start:
                for row_idx in range(start, r):
                    for c_idx in [1,2]:
                        cell = ws.cell(row_idx, c_idx)
                        cell.border = Border(
                            top=thick if row_idx==start else cell.border.top,
                            bottom=thick if row_idx==r-1 else cell.border.bottom,
                            left=thick if c_idx==1 else cell.border.left,
                            right=thick if c_idx==2 else cell.border.right
                        )
            start = r

def main(url):
    # domain global variable must be set here for check_email_security_txt to access it via nonlocal scope or passed
    global domain 
    domain = url.replace("http://", "").replace("https://", "").replace("www.", "").split('/')[0]
    
    # Version 52.1 (Syntax Fix & Forced Discovery & Reprompt Loop)
    print(f"\n🔍 INITIATING ANALYSIS FOR: {domain} (v52.1 - Syntax Fix & Forced Discovery & Reprompt Loop)\n") 

    base_dir = Path(__file__).resolve().parent
    dns_file_path = base_dir / f"{domain}_dns.txt"
    webservers_file_path = base_dir / f"{domain}_webservers.txt"
    
    # -------------------------------------------------------------------------
    # 1. BASE INFORMATION GATHERING (Kept outside loop as it's static)
    # -------------------------------------------------------------------------
    c_ip = "Unresolved"
    try: c_ip = socket.gethostbyname(domain)
    except: pass
    
    current_data_base = []; 
    reg, mail = "Hidden", "-"
    try:
        with SuppressStderr():
            w = whois.whois(domain)
            reg = clean_val(w.registrar)
            if w.emails: mail = str(w.emails[0] if isinstance(w.emails, list) else w.emails)
    except: pass
    geo = get_ip_details(c_ip)
    
    current_data_base.append(["Domain", domain]); current_data_base.append(["IP", c_ip])
    if geo and geo.get('status')=='success':
        current_data_base.append(["Location", f"{geo['city']}/{geo['country']}"])
        current_data_base.append(["ISP", geo['isp']]); current_data_base.append(["ASN", geo['as']])
    current_data_base.append(["Registrar", reg]); current_data_base.append(["Contact", mail])
    
    # --- START MAIN EXECUTION LOOP ---
    initial_discovery_run = False
    
    while True:
        # Reset data containers and flags for the new session/loop
        run_discovery = False
        run_portscan = False
        run_webperf = False
        run_mail_test = False
        
        subdomain_data_excel = []; root_dns_records_for_file = []; dns_display_list = []; scan_data = []; subdomain_data_console = []
        mail_test_results = [] 
        all_perf_data = [] # Moved initialization here
        
        # -------------------------------------------------------------------------
        # 2. MENU AND SELECTION
        # -------------------------------------------------------------------------
        
        if dns_file_path.exists():
            
            print("\n--- 💾 DNS Record File Found ---")
            print(f"File '{dns_file_path.name}' exists. Please select the action you want to perform:")
            print("1 - UPDATE DNS/Subdomain Records (Discovery)")
            print("2 - RUN Port Scan and Web Performance Test (Using existing DNS data)")
            print("3 - RUN ALL TESTS (Discovery + Port Scan + Web Perf)")
            print("4 - RUN Mail Tests (SMTP Health, Reputation, Email Auth)")
            choice = input("Your Selection (1/2/3/4): ").strip()
            
            if choice == '1':
                run_discovery = True
            elif choice == '2':
                # Option 2 selected: Load DNS data, then check webservers.txt for sub-menu
                print("⏳ Existing DNS and Subdomain records are loading from file...")
                subdomain_data_excel, root_dns_records_for_file, dns_display_list, scan_ips, subdomain_data_console = read_dns_data_from_file(domain, c_ip)
                
                if not scan_ips and (c_ip == "Unresolved" or not is_valid_ip(c_ip)):
                    print("\n[!] CRITICAL ERROR: Could not load any IP addresses from the existing DNS file. Cannot run Port Scan/Web Perf. Exiting.")
                    break
                    
                if webservers_file_path.exists():
                    print("\n--- 🌐 Webservers File Found ---")
                    print(f"File '{webservers_file_path.name}' exists. Please select the action for tests:")
                    print("1 - UPDATE Port Scan only (Skips Web Performance)")
                    print("2 - RUN Web Performance Test only (Uses existing Webservers list)")
                    print("3 - RUN BOTH Port Scan and Web Performance Test")
                    sub_choice = input("Your Selection (1/2/3): ").strip()
                    
                    if sub_choice == '1': run_portscan = True; run_webperf = False
                    elif sub_choice == '2': run_portscan = False; run_webperf = True
                    elif sub_choice == '3': run_portscan = True; run_webperf = True
                    else: print("[!] Invalid selection. Exiting."); break
                else:
                    print("\n[!] Webservers File Not Found. Running Port Scan and Web Performance Test...")
                    run_portscan = True; run_webperf = True

            elif choice == '3':
                run_discovery = True; run_portscan = True; run_webperf = True
            
            elif choice == '4':
                run_mail_test = True
                print("⏳ Loading existing DNS records to identify MX targets...")
                subdomain_data_excel, root_dns_records_for_file, dns_display_list, scan_ips, subdomain_data_console = read_dns_data_from_file(domain, c_ip)

                print("\n--- 📧 Mail Test Module ---")
                print("Please select the mail test action:")
                print("1 - Run SMTP Health Check (MXToolbox Simulation)")
                print("2 - Run Reputation Checks only (DNSBL, AbuseIPDB)")
                print("3 - Run ALL Mail Tests (SMTP + Reputation + Email Auth)")
                print("4 - Run Email Auth Checks only (SPF, DMARC, DKIM)") 
                mail_choice = input("Your Selection (1/2/3/4): ").strip()
                
                if mail_choice in ['1', '2', '3', '4']:
                    mail_test_results = run_mail_tests(domain, root_dns_records_for_file, mail_choice, c_ip)
                else: print("[!] Invalid selection. Exiting."); break

            else:
                print("[!] Invalid selection. Exiting."); break

        else:
            # --- FILE NOT FOUND: FORCED DISCOVERY AND PORTSCAN ---
            print("\n[!] DNS Record File Not Found. Running Auto-Discovery and Port Scan...")
            initial_discovery_run = True
            run_discovery = True
            run_portscan = True
            run_webperf = False # Skip webperf initially
            
        # -------------------------------------------------------------------------
        # 3. EXECUTION PHASE
        # -------------------------------------------------------------------------
        
        # --- DISCOVERY & DATA LOADING ---
        if run_discovery:
            subdomain_data_excel, root_dns_records_for_file, dns_display_list, scan_ips, subdomain_data_console = perform_discovery_tasks(domain, c_ip)
        
        # --- PORT SCAN ---
        if run_portscan:
            scan_data = run_portscan_tasks(scan_ips)
        elif not run_discovery:
            scan_data = ["Result", "Skipped by user selection.", "-"]
            
        # --- WRITE TXT FILES ---
        write_dns_flag = run_discovery 
        write_webservers_flag = run_portscan 
        
        if write_dns_flag or write_webservers_flag:
            write_domain_info_to_files(
                domain, c_ip, root_dns_records_for_file, subdomain_data_excel, scan_data if run_portscan else [],
                write_dns=write_dns_flag, write_webservers=write_webservers_flag
            )

        # -------------------------------------------------------------------------
        # 4. REPORTING AND LOOP CONTROL
        # -------------------------------------------------------------------------
        
        # Only proceed to reporting if a test was actually initiated
        if run_discovery or run_portscan or run_webperf or run_mail_test:
            
            # --- WEB PERF (Only runs if explicitly chosen via 2 or 3) ---
            if run_webperf:
                # This complex section remains as is, checking webservers.txt and running PSI
                urls_to_test = [f"https://{domain}"] 
                if webservers_file_path.exists():
                    try:
                        with open(webservers_file_path, 'r', encoding='utf-8') as f:
                            urls_from_webservers = []
                            for line in f:
                                if line.startswith('#'): continue
                                parts = line.split('|')
                                if len(parts) >= 1: urls_from_webservers.append(f"https://{parts[0].strip()}")
                        urls_to_test = list(set(urls_from_webservers)) 
                    except Exception as e: print(f"   [!] ERROR: Failed to read Webserver file: {e}. Only root domain will be tested.")
                
                print(f"⏳ Checking Web Performance for {len(urls_to_test)} total URL(s)...")
                all_perf_data = []
                for url_to_test in urls_to_test:
                    perf_result = None; error = None; MAX_RETRY = 3
                    for attempt in range(1, MAX_RETRY + 1):
                        print(f"   > Running Lighthouse on {url_to_test} (Attempt {attempt}/{MAX_RETRY})...")
                        perf_result, error = analyze_pagespeed(url_to_test)
                        if perf_result: print(f"   [SUCCESS] Lighthouse analysis successful on attempt {attempt}."); break 
                        elif attempt < MAX_RETRY: time.sleep(60) 
                    if perf_result: all_perf_data.append(perf_result)
                    elif error: all_perf_data.append({'url': url_to_test, 'error': error[0]})

            # --- HISTORY GATHERING (Always runs for complete report) ---
            print("⏳ Fetching History...")
            hist = get_rapiddns_history(domain)
            h_seen = set(); h_seen.add(c_ip); cnt=0
            history_excel = []; history_console = []
            if hist:
                for i, r in enumerate(hist):
                    if r['ip'] in h_seen: continue
                    h_seen.add(r['ip'])
                    sys.stdout.write(f"\r>> Processing: {r['ip']} ({i+1}/{len(hist)})  "); sys.stdout.flush()
                    h_geo = get_ip_details(r['ip'])
                    history_excel.append([f"--- RECORD #{cnt+1} ---", f"Date: {r['date']}"]); history_excel.append(["IP", r['ip']])
                    loc_str = "-"; isp_str = "-"
                    if h_geo and h_geo.get('status')=='success':
                        loc_str = f"{h_geo['city']}/{h_geo['country']}"; isp_str = h_geo['isp']
                        history_excel.append(["Loc", loc_str]); history_excel.append(["ISP", isp_str])
                    else: history_excel.append(["Info", "No Geo Data"])
                    history_console.append([r['date'], r['ip'], loc_str, isp_str[:30]]); cnt+=1
            if cnt==0: history_excel.append(["History", "No changes"]); history_console.append(["-", "No changes", "-", "-"])
            
            # --- REPORT GENERATION (CONSOL & EXCEL) ---
            print("\n\n✅ Analysis Complete. Printing Results...\n")
            
            # --- Consol output data preparation for unified display ---
            perf_console_data = []
            if all_perf_data:
                for item in all_perf_data:
                    url = item['url'].replace('https://', '')
                    if 'error' in item: perf_console_data.append([url, "N/A", "N/A", "N/A", item['error']])
                    else:
                        scores = item['scores']; metrics = item['metrics']
                        perf_console_data.append([url, scores.get('Performance', '-'), scores.get('Accessibility', '-'), scores.get('SEO', '-'), metrics.get('LCP', '-')])
            elif not run_webperf: perf_console_data.append(["N/A", "N/A", "N/A", "N/A", "Skipped by user selection."])
            else: perf_console_data.append(["N/A", "N/A", "N/A", "N/A", "Test failed or produced no data."])
            
            # --- DISPLAY 7 TABLES ---
            print("\n1️⃣  CURRENT INTELLIGENCE"); print(tabulate(current_data_base, headers=["PARAMETER", "VALUE"], tablefmt="fancy_grid")); print("---")
            print("\n2️⃣  DNS RECORDS (Root Domain)"); print(tabulate(dns_display_list, headers=["TYPE", "VALUE"], tablefmt="fancy_grid")); print("---")
            print("\n3️⃣  SUBDOMAIN DISCOVERY (Sublist3r)"); print(tabulate(subdomain_data_console, headers=["SUBDOMAIN", "TYPE", "VALUE/IP", "SOURCE"], tablefmt="fancy_grid")); print("---")
            
            print("\n4️⃣  PORT SCAN RESULTS")
            if isinstance(scan_data, list) and scan_data and isinstance(scan_data[0], dict):
                scan_display = [[r['ip'], f"{r['port']} ({r['service']})", r['banner']] for r in scan_data]
            else: scan_display = [scan_data]
            print(tabulate(scan_display, headers=["IP", "PORT", "BANNER"], tablefmt="fancy_grid")); print("---")

            print("\n5️⃣  WEB PERFORMANCE SUMMARY (Google Lighthouse - Selected URLs)"); 
            print(tabulate(perf_console_data, headers=["URL", "Perf Score", "Access.", "SEO Score", "LCP / Error"], tablefmt="fancy_grid")); print("---")
            
            mail_console_data = []
            if mail_test_results:
                for r in mail_test_results:
                    target = r.get('Target'); test = r.get('Test'); status = r.get('Status'); details = r.get('Details'); source = r.get('Source', 'N/A')
                    color_status = status; symbol = ""
                    if status in ['PASS', 'CLEAN', 'FOUND (Valid)']: color_status = f"{GREEN}{status}{ENDC}"; symbol = CHECK
                    elif status in ['FAIL', 'LISTED', 'MISSING', 'INVALID', 'API_ERROR']: color_status = f"{RED}{status}{ENDC}"; symbol = CROSS
                    elif status in ['SKIP', 'N/A', 'ERROR']: color_status = f"{status}"; symbol = "➖"
                    mail_console_data.append([symbol, target, test, color_status, source, details])
            elif run_mail_test: mail_console_data.append(["➖", "N/A", "N/A", "N/A", "N/A", "No MX records found or test skipped."])
            else: mail_console_data.append(["➖", "N/A", "N/A", "N/A", "N/A", "Skipped by user selection."])

            print("\n6️⃣  MAIL TEST RESULTS (SMTP Health / DNSBL / AbuseIPDB / Email Auth)");
            print(tabulate(mail_console_data, headers=["", "TARGET", "TEST TYPE", "STATUS", "SOURCE", "DETAILS"], tablefmt="fancy_grid")); print("---")

            print("\n7️⃣  HOSTING / A-RECORD HISTORY (DNS History)"); print(tabulate(history_console, headers=["DATE", "OLD IP ADDRESS", "LOCATION", "ISP"], tablefmt="fancy_grid")); print("---")
            
            # --- EXCEL EXPORT ---
            # Prepare DataFrames for Excel (using existing logic)
            perf_rows = []
            for result in all_perf_data:
                url = result['url']
                if 'error' in result: perf_rows.append([url, "Error", result['error'], "-", "-", "-", "-", "-"])
                else:
                    scores = result['scores']; metrics = result['metrics']
                    perf_rows.append([url, scores.get('Performance'), scores.get('Accessibility'), scores.get('SEO'), metrics.get('FCP'), metrics.get('LCP'), metrics.get('Speed Idx'), ' | '.join([f"{r[0]} ({r[1]})" for r in result['recommendations'][:3]])])

            perf_df = pd.DataFrame(perf_rows, columns=["URL", "Performance", "Accessibility", "SEO", "FCP", "LCP", "Speed Index", "Key Recommendations"])
            history_df = pd.DataFrame(history_excel, columns=["PARAM","VAL"])
            current_df = pd.DataFrame(current_data_base, columns=["PARAM","VAL"])
            dns_df = pd.DataFrame(dns_display_list, columns=["TYPE","VAL"])
            
            mail_excel_data = []
            for r in mail_test_results:
                row = {'Target': r['Target'], 'Test': r['Test'], 'Status': r['Status'], 'Details': r['Details'], 'Source': r.get('Source', '-')}
                if 'URL' in r: row['URL'] = r['URL']
                else: row['URL'] = "-"
                mail_excel_data.append(row)

            mail_df = pd.DataFrame(mail_excel_data, columns=['Target', 'Test', 'Status', 'Details', 'Source', 'URL'])
            if 'URL' in mail_df.columns:
                mail_df['URL'] = mail_df['URL'].apply(lambda x: f'=HYPERLINK("{x}","Link")' if x.startswith("http") else "-") 
            
            if isinstance(scan_data, list) and scan_data and isinstance(scan_data[0], dict):
                 scan_df = pd.DataFrame([[r['ip'], f"{r['port']} ({r['service']})", r['banner']] for r in scan_data], columns=["IP","PORT","BANNER"])
            else:
                 scan_df = pd.DataFrame([["N/A", "N/A", scan_data[1] if isinstance(scan_data, list) and len(scan_data)>1 else "Skipped/No Data"]], columns=["IP","PORT","BANNER"])

            sub_df = pd.DataFrame([[s['Subdomain'], s['Type'], s['Value/IP'], s['Last_Seen']] for s in subdomain_data_excel], columns=["Subdomain", "Type", "Value/IP", "Last_Seen"])
            
            sheet_map = {'Current': current_df, 'DNS_Root': dns_df, 'Ports': scan_df, 'Web_Perf': perf_df, 'Subdomains': sub_df, 'History': history_df, 'Mail_Tests': mail_df }
            
            try:
                ask = input(f"\n💾 Save report to '{domain}.xlsx'? (Y/N): ").lower()
                if ask in ['y','yes']:
                    fname = f"{domain}.xlsx"
                    file_exists = os.path.exists(fname)
                    
                    writer_mode = 'a' if file_exists else 'w'
                    writer_kwargs = {'engine': 'openpyxl', 'mode': writer_mode, 'if_sheet_exists': 'replace'}
                    print(f"[i] {'Existing' if file_exists else 'Creating new'} Excel file. Updating sheets...")

                    with pd.ExcelWriter(fname, **writer_kwargs) as writer:
                        for sheet_name, df in sheet_map.items(): df.to_excel(writer, sheet_name=sheet_name, index=False)
                        wb = writer.book
                        for sn in writer.sheets:
                            ws = writer.sheets[sn]
                            for c in ws[1]: c.font = Font(bold=True)
                            for r in ws.iter_rows(min_row=2, max_col=1): 
                                for c in r: c.font = Font(bold=True)
                            ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 60
                        
                        if 'Web_Perf' in writer.sheets:
                            ws_perf = writer.sheets['Web_Perf']; red = PatternFill(start_color='FFC7CE', fill_type='solid'); green = PatternFill(start_color='C6EFCE', fill_type='solid')
                            for row in ws_perf.iter_rows(min_row=2, min_col=2, max_col=4):
                                for cell in row:
                                    try: 
                                        val = int(cell.value)
                                        if val < 50: cell.fill = red
                                        elif val > 89: cell.fill = green
                                    except: 
                                        pass # <<< SYNTAX HATA DÜZELTMESİ BURADA <<<
                        if 'History' in writer.sheets: style_history_sheet(writer.sheets['History'])
                        if 'Subdomains' in writer.sheets: style_subdomain_sheet(writer.sheets['Subdomains']) 
                        
                        if 'Mail_Tests' in writer.sheets:
                            ws_mail = writer.sheets['Mail_Tests']
                            if ws_mail.max_column >= 4: ws_mail.column_dimensions['D'].width = 80 
                            for row in ws_mail.iter_rows(min_row=2, min_col=3, max_col=3):
                                for cell in row:
                                    if cell.value in ['PASS', 'CLEAN', 'FOUND (Valid)']: cell.fill = green
                                    elif cell.value in ['FAIL', 'LISTED', 'MISSING', 'INVALID', 'API_ERROR'] or (isinstance(cell.value, str) and cell.value.startswith('LISTED')): cell.fill = red

                    print(f"\n[+] SUCCESS: Report saved/updated as '{fname}'.")
                else: print("[-] Cancelled")
            except Exception as e: print(f"[!] ERROR saving Excel: {e}")

            # --- LOOP CONTROL ---
            if initial_discovery_run:
                print(f"\n--- 📝 Initial Data Collection Complete ---")
                print(f"Files '{dns_file_path.name}' and '{webservers_file_path.name}' have been created.")
                print(f"Now restarting to present the full interactive menu (based on new data).")
                initial_discovery_run = False # Reset flag if we were inside the logic.
                continue # Loop back to the beginning to show the full menu
            
            # After a successful manual run, prompt to exit or continue
            final_exit = input("\n[?] Analysis complete. Run another test or exit? (Type 'exit' or press Enter to continue): ").strip().lower()
            if final_exit == 'exit':
                break
        
        else: # If user input was invalid (e.g., in menu 1/2/3/4)
            break
            
if __name__ == "__main__":
    if len(sys.argv)>1: main(sys.argv[1])
    else: main(input("Enter target domain (e.g., google.com): "))